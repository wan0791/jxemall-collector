"""
Excel 导出模块 — 按模板 13 列格式 + 超链接 + 电脑高亮
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import db, config
# DB 连接由 app.py 的 _init() 统一初始化，exporter 运行时 pool 已就绪

COLUMNS =[
    "品类", "品牌", "XC", "产品名称型号", "最高限价", "成交价",
    "成交总价", "数量", "地区", "供货商", "采购单位",
    "成交时间", "是否集采"
]

CITY_SHEET_NAMES = {
    "南昌": "1南昌", "九江": "2九江",
    "上饶": "3上饶", "鹰潭": "4鹰潭", "景德镇": "5景德镇",
    "宜春": "6宜春", "萍乡": "7萍乡",
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
COMP_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
LINK_FONT = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
THIN_BDR = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin')
)

def _init_sheet(ws):
    for i, cn in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=i, value=cn)
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, THIN_BDR
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # 列宽
    widths = {
        "品类": 8, "品牌": 12, "XC": 6, "产品名称型号": 40,
        "最高限价": 12, "成交价": 12, "成交总价": 12, "数量": 8,
        "地区": 10, "供货商": 28, "采购单位": 28,
        "成交时间": 14, "是否集采": 10,
    }
    for i, cn in enumerate(COLUMNS, 1):
        if cn in widths:
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = widths[cn]
    ws.auto_filter.ref = ws.dimensions
    # 冻结首行
    ws.freeze_panes = 'A2'


def _append_row(ws, rd):
    rn = ws.max_row + 1
    row_data = [
        rd.get("category", ""),
        rd.get("brand", ""),
        "",  # XC
        f"{rd.get('item_name', '')} / {rd.get('item_spec', '')}".strip(" /"),
        "",  # 最高限价
        rd.get("item_unit_price") or "",
        rd.get("item_total_price") or "",
        rd.get("item_qty") or "",
        rd.get("city_name", ""),
        rd.get("supplier", ""),
        rd.get("buyer", ""),
        rd.get("released_at", ""),
        "",  # 是否集采
    ]

    is_comp = rd.get("is_computer", 0)

    for i, val in enumerate(row_data, 1):
        c = ws.cell(row=rn, column=i)
        c.value = val
        c.font = BODY_FONT
        c.border = THIN_BDR
        if i == 12:  # 成交时间居中
            c.alignment = Alignment(horizontal='center', vertical='center')
        elif i in (5, 6, 7, 8):  # 金额/数量右对齐
            c.alignment = Alignment(horizontal='right', vertical='center')
        elif i == 4:  # 产品名称型号
            c.alignment = Alignment(wrap_text=True, vertical='center')

    # 电脑行高亮
    if is_comp:
        for i in range(1, len(COLUMNS) + 1):
            ws.cell(row=rn, column=i).fill = COMP_FILL

    # 超链接列 — 在标题列后面附加
    detail_url = rd.get("detail_url", "")
    # 我们在第 14 列（隐藏）放超链接
    link_col = len(COLUMNS) + 1
    if detail_url:
        ws.cell(row=rn, column=link_col).value = f'=HYPERLINK("{detail_url}","查看原文")'
        ws.cell(row=rn, column=link_col).font = LINK_FONT
        ws.cell(row=rn, column=link_col).alignment = Alignment(vertical='center')


def export(date_from=None, date_to=None, city=None, computer_only=False, mark=None):
    """导出 Excel"""
    rows, total, _ = db.query_contracts(
        date_from, date_to, city, computer_only, mark=mark, limit=100000
    )
    if not rows:
        return None

    wb = Workbook()
    wb.remove(wb.active)

    # 按城市分组
    city_data = {}
    for rd in rows:
        cn = rd.get("city_name", "其他")
        if cn not in CITY_SHEET_NAMES:
            cn = "其他"
        if cn not in city_data:
            city_data[cn] = []
        city_data[cn].append(rd)

    for cn, sheet_name in CITY_SHEET_NAMES.items():
        data = city_data.get(cn, [])
        if not data:
            continue
        ws = wb.create_sheet(sheet_name)
        _init_sheet(ws)
        for rd in data:
            _append_row(ws, rd)

    # 其他城市
    if "其他" in city_data:
        ws = wb.create_sheet("其他城市")
        _init_sheet(ws)
        for rd in city_data["其他"]:
            _append_row(ws, rd)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(config.ROOT, f"电子卖场合同_{ts}.xlsx")
    wb.save(path)
    return path
